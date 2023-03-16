import openpyxl
from bigbear_mysql import dosql
from bigbear_ini import moren,rid,ghid,qjltoken
from requests_toolbelt.multipart.encoder import MultipartEncoder
from bigbear_wx import getinfo,get_gp_nickname,sdfile,sdtxt
from itertools import combinations
from bigbear_base import daymonth,chinesetoupper,string_similar,logger
from pathlib import Path
from bigbear_web_api import get_company_name
from bigbear_cloud_api import updateexp
import jionlp as jio
import os
import requests
import  time
import json
import pythoncom
from win32com.client import DispatchEx

rows4=0
columns4=0
marke4=0


def get_excel_val(ws,row,column):
    print("当前列：%d"%(column))
    if column>0:
        return ws.cell(row=row,column=column).value
    else:
        return "无此列信息"
def file_analy_agent(filepath,fromgp):
    nick=''
    sql="select * from 群文件配置表 where id='"+fromgp+"'"
    ret=dosql(sql)
    has_gp=False
    #获取表头
    print(len(ret))
    if len(ret)>0:
        #如果查询表头没有配置,则使用默认的
        biaotou=ret[0]['唯一商品名表头']
        has_gp=True
        commonlist=['SKU名称','商品名称']
        if biaotou in commonlist:
            biaotou=commonlist
    else:
        #默认表格信息为list,多元素数组
        biaotou=['SKU名称','商品名称']
    marke4=-1
    wb=openpyxl.load_workbook(filepath)
    global  sheet4excel
    sheet4excel=wb[wb.sheetnames[0]]
    sheetcount=len(wb.sheetnames)
    columns4=sheet4excel.max_column
    rows4=del_blank_rows(filepath)

    for i in range(1, columns4 + 1):
        cellval = sheet4excel.cell(row=1,column=i).value
        #判断表头是否为list
        if type(biaotou)==list:
            for i_b in biaotou:
                if cellval is not None:
                    if i_b in cellval:
                        # 标记运单列
                        marke4 = i
        else:
            print("指定了表头")
            est=biaotou
            if cellval is not None:
                if est in cellval:
                    # 标记运单列
                    marke4 = i

    print("当前表头列在：{}".format(marke4))
    onlyset=set()
    for i in range(2, rows4 + 1):
        cellval = sheet4excel.cell(column=marke4,row=i).value
        onlyset.add(cellval)
    onlylist=list(onlyset)
    thelist=gettheset(onlylist)
    rowobj = {}
    if len(thelist)>1:
        for i in thelist:
            # 得到每个商品

            # print(i)
            for ia in range(2, rows4 + 1):
                cellval = sheet4excel.cell(column=marke4, row=ia).value
                if i in cellval:
                    if i in rowobj.keys():
                        rowobj[i].append(str(ia))
                    else:
                        rowobj[i] = []
                        rowobj[i].append(str(ia))

        print(rowobj)
        creat_init_excel(thelist, rowobj,fromgp)
    else:
        #当前为单产品则直接将其转发至对应群
        #通过商品名查询需要转发到的群
        itemval=sheet4excel.cell(column=marke4,row=2).value
        getname = getinfo(fromgp)
        nick=getname['result']['nickBrief']
        print("群昵称大写：%s"%(nick))
        newfilepath=r'C:\orderFiles'+"\\"+daymonth()+itemval+nick+".xlsx"
        print("原路径文件：%s,另存为：%s"%(filepath,newfilepath))
        wb.save(newfilepath)
        print('当前产品为单品')
        sql="select id from 商品列表 where item='%s'"%(itemval)
        ret=dosql(sql)

        #如果查不到来源则发送给默认接收人文件,并且提示默认接收人配置商品列表
        if len(ret)>0:
            sdfile(ret[0]['id'], filepath)
        else:
            sdtxt(moren,"当前商品%s未配置,请复制下条命令消息到对应群配置商品"%(itemval))
            sdtxt(moren,"@配置商品+%s"%(itemval))
            sdfile(moren,filepath)
        return
#得到唯一商品名称
def check_file_exists(mingzi,filename,obj):
    rootpath = "C:\orderFiles"+"\\"+filename
    my_excel_files = Path(rootpath)
    if my_excel_files.exists():
        #print('存在')
        pass
    else:
        newwb=openpyxl.load_workbook('C:\orderFiles\我是模板.xlsx')
        newwb.save(rootpath)
    nwb=openpyxl.load_workbook(rootpath)
    nsheet=nwb[nwb.sheetnames[0]]
    #初始化表格title
    for col, val in enumerate(sheet4excel['1'], start=1):
        nsheet.cell(row=1, column=col).value = val.value
    #写入所有内容
    #obj[mingzi]得到原来文件的所在行依次写入当前文件，len(obj[mingzi])
    startpos=2
    for i in range(0,len(obj[mingzi])):
        #当前在原来表格行obj[mingzi][i]
        for col, val in enumerate(sheet4excel[obj[mingzi][i]], start=1):
            nsheet.cell(row=startpos, column=col).value = val.value
        startpos=startpos+1
    nwb.save(rootpath)
    sdfile(moren,rootpath)
# 得到清洗后的每个商品所在的行数
def creat_init_excel(namelist,obj,gpname):
    for i in namelist:
        #获取当前的产品名，拼接文件名
        filename=daymonth()+i+chinesetoupper(gpname)+".xlsx"
        check_file_exists(i,filename,obj)

def del_blank_rows(path):
    blank_count = 0
    if os.path.exists(path):
        # 如果存在
        wb = openpyxl.load_workbook(path)
        sheet1 = wb[wb.sheetnames[0]]
        max_row = sheet1.max_row
        max_column = sheet1.max_column
        # print(max_row)
        # print(max_column)
        # 开始循环,将每行大部分为空的行数从总行数移除
        # 循环每行
        for row in range(1, max_row + 1):
            # 循环每列
            # print("当前行号：{}".format(row))
            for column in range(1, max_column + 1):
                values = sheet1.cell(row=row, column=column).value
                # print(values)
                # 如果为空或者没值
                if values == "" or values is None:
                    blank_count = blank_count + 1

            if blank_count > max_column // 2:
                max_row = max_row - 1
                # print("当前行{0},当前列{1}".format(row, column))
            blank_count = 0
        # print('循环完后maxrow还有多少行')
        # print(max_row)
        return max_row
    else:
        return "当前文件不存在请确认,原文件路径为：%s"%(path)

def gettheset(onlylist):
    listcopy = onlylist

    for i in combinations(listcopy, 2):
        if len(onlylist)>2:
            res = string_similar(i[0], i[1])
            if res > 0.7:
                if len(i[0]) > len(i[1]):
                    onlylist.pop(onlylist.index(i[0]))
                    print(onlylist)
                elif len(i[0]) < len(i[1]):
                    onlylist.pop(onlylist.index(i[1]))
                    print(onlylist)
                else:
                    onlylist.pop(onlylist.index(i[0]))
                    print(onlylist)
                    # 添加一个列表后补
                    etext = jio.extract_chinese(i[0])
                    ett = etext[0]
                    if len(etext) > 1:
                        ett = max(etext, key=len, default='')
                    onlylist.append(ett)
                gettheset(onlylist)
            else:
                onlylist=onlylist
        else:
            res = string_similar(onlylist[0], onlylist[1])
            if res > 0.7:
                if len(onlylist[0]) > len(onlylist[1]):
                    onlylist.pop(onlylist.index(onlylist[0]))
                    print(onlylist)
                elif len(onlylist[0]) < len(onlylist[1]):
                    onlylist.pop(onlylist.index(onlylist[1]))
                    print(onlylist)
                else:
                    onlylist.pop(onlylist.index(onlylist[0]))
                    etext = jio.extract_chinese(onlylist[0])
                    ett = etext[0]
                    if len(etext) > 1:
                        ett = max(etext, key=len, default='')
                    onlylist[0]=ett
                    print(onlylist)

            return onlylist
    return onlylist

def tongbushuju_gp(file,fromwho):

    wb = openpyxl.load_workbook(file)
    st1 = wb[wb.sheetnames[0]]
    mrow = st1.max_row
    mcol = st1.max_column
    # 数量
    markn = -1
    # 规格
    markg = -1
    #订单号列
    marko=-1
    # 金额
    markcash = -1
    # 序列号
    marknl = -1
    # 收货人
    markname = -1
    # 手机号
    markphone = -1
    # 下单时间
    marktime = -1
    # 下单地址
    markaddress = -1
    # 商品编码
    markcode = -1
    # 团长ID
    markID = -1
    #商品名称
    markitem=-1

    ordertimelist=["下单时间"]
    specificationslist=["规格", "SKU规格","规格名称"]
    addlist=["详细地址", "收货地址", "完整地址","收件人完整地址"]
    snolist=["下单序号", "跟团号 / 下单序号", "跟团号"]
    consigneelit=["收件人", "收货人", "收件人姓名"]
    quantitieslist = ["数量", "SKU数量"]
    itemlist = ["商品名称", "SKU名称"]
    phonelist = ["联系方式", "联系电话", "收件人手机号码", "电话"]
    elist = ["快递单号", "物流单号", "运单号"]
    clist = ["快递公司", "物流公司"]
    olist = ["订单号", "订单编号", "商城订单号","订单号(勿删)"]
    itemcode=["商品编码","SKU编号","订单商品号(勿删)"]
    alreadylist=[]
    for im in range(1, mcol + 1):
        print("当前检测列为：%d"%(im))
        cellval = st1.cell(row=1, column=im).value
        print("当前列内容为：%s" % (cellval))
        # 下单序号	接龙号	收货人	联系电话	商品名称	商品编码	规格	数量	商品金额	订单总金额	下单时间	省	市	收货地址	团长ID
        #下单序号
        print("开始判断下单序号")
        if cellval in snolist:
            marknl = im
            continue
        #判断订单号列
        print("开始判断订单号列")
        for i in olist:
            if cellval==i:
                marko=im
                break
        #判断商品名称
        if cellval in itemlist:
            markitem = im
            continue
        #收货人
        print("开始判断收货人")
        if cellval in consigneelit:
            markname = im
            continue
        #联系方式
        print("开始判断联系方式")
        if cellval in phonelist:
            markphone = im
            continue
        #规格
        print("开始判断规格")
        for i in specificationslist:
            if cellval == i:
                markg = im
                break

        print("开始判断数量")

        if cellval in quantitieslist:
            markn = im
            continue
        print("开始判断订单总金额")

        if cellval == '订单总金额':
            markcash = im
            continue
        print("开始判断下单时间")
        if cellval in ordertimelist:
            marktime = im
            continue
        print("开始判断收货地址")
        for i in addlist:
            if cellval == i:
                markaddress = im
                break

        print("开始判断收货地址")
        if cellval == '团长ID':
            markID = im
            continue
        print("开始判断商品编码")
        if cellval in itemcode:
            markcode = im
    #获取群昵称
    nickname=get_gp_nickname('nick',fromwho)
    #开始获取数据并上传同步到服务器
    mrow= del_blank_rows(file)
    for i in range(2, mrow+1):
        print('开始同步表格订单信息到服务器')
        ordernumb = get_excel_val(st1,i,marko)
        if ordernumb != "无此列信息" or ordernumb is not None:
            print('初始化当前规格信息')
            ordertime = get_excel_val(st1,i,marktime)
            phonenum =  get_excel_val(st1,i,markphone)
            consignee = get_excel_val(st1,i,markname)
            serialnum = get_excel_val(st1,i,marknl)
            address = get_excel_val(st1,i,markaddress)
            specifications = get_excel_val(st1,i,markg)
            quantities = get_excel_val(st1,i,markn)
            totalamount = get_excel_val(st1,i,markcash)
            headid = get_excel_val(st1,i,markID)
            codenum = get_excel_val(st1,i,markcode)
            itemname=get_excel_val(st1,i,markitem)
            # 说明当前有数据
            print("订单总数：%d,订单号:%s：" % (mrow - 2, ordernumb))
            # 更新数据库数据进行对比
            dizhi = 'https://91e787c8-df7a-42ad-b4fa-12dde982b53d.bspapp.com/infos'
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.163 Safari/537.36',
                'content-type': 'application/json'
            }
            data={
                                 "date": daymonth(),
                                 "itemName": itemname,
                                 "orderNo": ordernumb,
                                 "orderTime": ordertime,
                                 "phone": phonenum,
                                 "consignee": consignee,
                                 "serialNum": serialnum,
                                 "address": address,
                                 "specifications": specifications,
                                 "quantities": quantities,
                                 "totalAmount": totalamount,
                                 "headID": headid,
                                 "codeNum": codenum,
                                 'status': 0,
                                 'expNo':[],
                                 "fromwhere": fromwho,
                                "nick":nickname
                             }
            data=str(data)
            data=data.encode('utf-8')
            r = requests.post(dizhi, headers=headers,
                             data=data)
            # 获取反馈信息,如果存在则对单元格进行删除操作
            # r.json()  类型为dict
            print(r.json())
            rdict = r.json()
            # 判断是否已经发货过通过上传订单号获取反馈
            if 'id' in rdict:
                # 如果未发货则存在
                print('当前订单号首次出现')
            else:
                print("当前订单已存在%s"%(ordernumb))


        else:
            print('订单为空')
            sdtxt(fromwho,"当前表格存在订单号信息为空，无法完成同步，请校验表格是否正确")
        # 完成订单检测后再对当前未回单的群进行催单

        time.sleep(0.5)
    wb.close()

#检测上传文件的
def check_upload_excel(fromwho,filename,path,sendby):
    #快递单数总计
    express_total=0
    #订单总计
    order_total=0
    #等待十秒,目的是希望它能自动下载完了
    time.sleep(1)
    #快递单号,快递公司,订单号所在列
    e_column=company_column=order_column=-1
    #判断是否存在此文件
    #判断单元格内容是否为空的计次结果
    blank_count=0
    if os.path.exists(path):
        #如果存在
        wb=openpyxl.load_workbook(path)
        sheet1=wb[wb.sheetnames[0]]
        max_row=sheet1.max_row
        max_column=sheet1.max_column
        #print(max_row)
        #print(max_column)
        #开始循环,将每行大部分为空的行数从总行数移除
        #循环每行
        for row in range(1,max_row+1):
            #循环每列
            #print("当前行号：{}".format(row))
            for column in range(1,max_column+1):
                values=sheet1.cell(row=row,column=column).value
                #print(values)
                #如果为空或者没值
                if values=="" or values is None:
                    blank_count=blank_count+1

            if blank_count>max_column//2:
                max_row=max_row-1
                #print("当前行{0},当前列{1}".format(row, column))
            blank_count=0
        #print('循环完后maxrow还有多少行')
        #print(max_row)
        #循环表头所在列：
        elist={"快递单号","物流单号","运单号"}
        clist={"快递公司","物流公司"}
        olist={"订单号","订单编号"}
        #对结果进行重复比较
        eset=set()
        cset=set()
        oset=set()
        del_col=[]


        for column in range(1,max_column+1):
            #快递单号
            for i in elist:
                values=sheet1.cell(row=1,column=column).value
                if i in values:
                    e_column=column
                    eset.add(column)
            #快递公司
            for i in clist:
                values=sheet1.cell(row=1,column=column).value
                if i in values:
                    company_column=column
                    cset.add(column)
            #订单号所在列
            for i in olist:
                values=sheet1.cell(row=1,column=column).value
                if i in values:
                    order_column=column
                    oset.add(column)
        #输出所在列进行判断

        if max_row>1:
            values=sheet1.cell(row=2,column=order_column).value
            #对value进行判断,如果当前订单来源于自身,则将订单进行进一步处理并上传,如果来自其他地方则直接转发
            datas={
                "mode":"checkOrderNo",
                "orderNo":values
            }
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.163 Safari/537.36',
                'content-type': 'application/json'
            }
            ret=requests.post("https://91e787c8-df7a-42ad-b4fa-12dde982b53d.bspapp.com/infos",headers=headers,data=json.dumps(datas))
            rjson=ret.json()
            if rjson['affectedDocs']>0:
                upbyid=rjson['data'][0]['fromwhere']
                if upbyid==rid:
                    print("继续执行,检测和补全信息")
                    #upbyid等于rid就继续执行,检测和补充单号
                else:
                    wb.close()
                    #sdfile and return
                    if "chatroom" in upbyid:
                        logger.info("当前文件:%s已经转发至%s"%(path,upbyid))
                        sdfile(upbyid,path)
                        sdtxt(fromwho,"当前回单文件已处理")
                        return
                    else:
                        #如果当前订单来源于其他机器人上传，则提示错误信息，但必须是发送文件的人，fromwho是群.sendby是发送人，为了避免2个机器人同时在群，导致另外个机器人进行错误操作
                        if fromwho==upbyid or upbyid== sendby:
                            logger.info("当前文件上传人和发送人相同，将忽略")
                            print("当前文件上传人，发送人相同，将忽略")
                            #sdtxt(moren,"当前文件来源错误，%s"%(path))
                            return

            else:
                sdtxt(moren,"当前订单未录入，来源错误，查看详情，请打开文件路径：%s"%(path))


        #print(e_column,company_column,order_column)
        if len(eset)>1:
            print("当前快递单号存在多列的情况")
            #对每列中内容就行判断,如果不等则将前者复制到后者,并且在最后删除列,总行数为删除空白行过后的行数,不会对超过现有行的进行处理
            eset=list(eset)
            for row in range(1,max_row+1):
                if len(eset)==2:
                    #开始对每个单元格进行判断
                    values=sheet1.cell(row=row,column=eset[0]).value
                    value=sheet1.cell(row=row,column=eset[1]).value
                    #如果不相等,且前者不等于空和不等于None赋值给后者
                    if values != value and values != "" and values is not None and value == "" or value is None:
                        sheet1.cell(row=row,column=eset[1]).value=values
                else:
                    #当前列已经超过2个了
                    sdtxt(fromwho,"当前文件,存在多个运单号列,无法自动处理。文件路径："+path)
            #添加删除当前的列
            del_col.append(eset[0])
        if len(cset)>1:
            print("当前快递公司存在多列的情况")
            cset=list(cset)
            for row in range(2,max_row+1):
                if len(cset)==2:
                    #开始对每个单元格进行判断
                    values=sheet1.cell(row=row,column=cset[0]).value
                    print(cset[0])
                    print(values)
                    value=sheet1.cell(row=row,column=cset[1]).value
                    print(value)
                    #如果不相等,且前者不等于空和不等于None赋值给后者
                    if values != value and values != "" and values is not None and value == "" or value is None:
                        sheet1.cell(row=row,column=cset[1]).value=values
                        print("复制成功")
                    else:
                        print("无需复制")
                else:
                    #当前列已经超过2个了
                    sdtxt(fromwho,"当前文件,存在多个物流公司列,无法自动处理。文件路径："+path)
            #添加删除列
            del_col.append(cset[0])
        if len(oset)>1:
            print("当前订单号存在多列的情况")
        #保存文档

        #循环删除当前列
        deltime=0
        if len(del_col)>0:
            print("当前存在需要删除的列")
            for i in del_col:
                if i - deltime > 1:
                    i = i - deltime
                else:
                    i = i
                sheet1.delete_cols(i)
                deltime = deltime + 1
        #读取当前各列数据
        e_column=e_column-deltime
        company_column=company_column-deltime
        order_column=order_column-deltime
        autocname=""
        ftime=True
        wb.save(path)
        time.sleep(1)
        nwb=openpyxl.load_workbook(path)
        nsheet1=nwb[nwb.sheetnames[0]]
        print(e_column,company_column,order_column) #删除列后当前的情况列,循环列中的空格并补全
        for row in range(2,max_row+1):
            e_val=nsheet1.cell(row=row,column=e_column).value
            if e_val is not None:
                express_total=express_total+1

            c_val=nsheet1.cell(row=row,column=company_column).value
            copyval=c_val
            #如果公司名称为空则考虑自动补全
            if  c_val is None and autocname=="":
                ret=get_company_name(e_val)
                if ftime:
                    if "数据来自" in ret:
                        #说明成功了
                        autocname=ret[4:]
                        nsheet1.cell(row=row, column=company_column).value=autocname
                        ftime=False
                    else:
                        #说明获取公司失败了
                        autocname="自动补全失败"
                        sdtxt(fromwho,"大笨熊助手自动补全物流公司失败：{}".format(path))
                else:
                    if autocname=="":
                        ret = get_company_name(e_val)
                        if "数据来自" in ret:
                            # 说明成功了
                            autocname = ret[4:]
                            nsheet1.cell(row=row, column=company_column).value = autocname
                    else:
                        print("当前有快递公司为空,自动补全中")
                        nsheet1.cell(row=row, column=company_column).value = autocname
            else:
            #说明公司名称不为空

                if ftime==False:
                    nsheet1.cell(row=row, column=company_column).value = autocname


            o_val=nsheet1.cell(row=row,column=order_column).value
            if o_val is not None:
                order_total=order_total+1
            #更新订单号物流信息到云服务器
            print(e_val,c_val,o_val)
            if e_val is not None and o_val is not None:
                updateexp(o_val,1,e_val)
            else:
                if o_val is not None:
                    updateexp(o_val,404)


        if order_total == express_total:
            pass
        else:
            #运单号和订单号总数不相同
            str="当前文件内，订单总计%d单，运单号总计%d单，由于缺少运单号，请打开文件查看具体原因"%(order_total,express_total)
            sdtxt(fromwho,str)
            logger.warning(str)
        #todo自动补全后开始上传到对应的平台
        logger.info("开始上传表格，更新物流信息，结果请查看群消息")
        print("开始上传订单")
        nwb.save(path)
        #尝试保存文件
        return uploadfiles(path,filename,fromwho)
    else:
        #回调函数,继续判断文件
        check_upload_excel(fromwho,filename,path,sendby)

#上传文件到群接龙
def uploadfiles(filepath,filename,fromgp):
    urlget='https://apipc.qunjielong.com/order-logistics/api/pc/logistics/get_upload_result?requestId='
    url ='https://apipc.qunjielong.com/order-logistics/api/pc/logistics/upload_pc_template/'+ghid+'/0'
    headers={
    'Accept': 'application/json, text/plain, */*',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Authorization': qjltoken,
    'Connection': 'keep-alive',
    'ghId': ghid,
    'Host': 'apipc.qunjielong.com',
    'Origin': 'https://pro.qunjielong.com',
    'Referer': 'https://pro.qunjielong.com/',
    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="102", "Google Chrome";v="102"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': "Windows",
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-site',
    'User-Agent':r'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
    }
    multipart_encoder = MultipartEncoder(
        fields={'file': (filename,open(filepath,'rb'),'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Expires': '0'}),
       },boundary='----WebKitFormBoundaryJ2aGzfsg35YqeT7X')
    headers['Content-Type'] = multipart_encoder.content_type
    result=requests.post(url,headers=headers,data=multipart_encoder)
    print("当前上传结果为")
    print(result.json())
    time.sleep(1)
    hget={
    'Host': 'apipc.qunjielong.com',
    'Connection': 'keep-alive',
    'Accept': '*/*',
    'Authorization': qjltoken,
    'Origin': 'https://pro.qunjielong.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-site',
    'Sec-Fetch-Dest': 'empty',
    'Referer': 'https://pro.qunjielong.com/',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'zh-CN,zh;q=0.9'
    }
    if result.json()['code']==200 and result.json()['success']:
        print(result.json())
        time.sleep(2)
        urlget=urlget+result.json()['data']
        r=requests.get(urlget,headers=hget)
        print(r.json())
        time.sleep(2)
        if r.json()['code']==200 and  r.json()['success']:
            #如果回单失败等于0，那就发送全部成功
            r = requests.get(urlget, headers=hget)
            print(r.json())
            time.sleep(1)
            if r.json()['data']['failItem']==0:
                geshu=r.json()['data']['successItem']
                #通知更新完成，并且准备删除文件
                sdtxt(fromgp,filename+'当前总计更新物流信息成功{}单！'.format(geshu))
                return True

            else:
                #通知失败
                fileurl="https://res0.shangshi360.com/"+r.json()['data']['fileName']
                msg= '当前文件存在回单失败订单，数量为' + str(r.json()['data']['failItem']) + '具体原因请查看表格链接:'+fileurl
                sdtxt(fromgp, filename + msg)
                return False
        else:
            sdtxt(fromgp,"回单失败，原因未知，请手动上传表格")
            return False

    else:
        retmsg=result.json()['msg']
        logger.info(retmsg)
        sdtxt(moren,retmsg+",请根据提示原因处理！")
